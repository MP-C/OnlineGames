### 1 Completar o exercício aqui
valor1 = float(input("Introduz um numero: "))
valor2 = float(input("Introduz outro numero: "))

# 1. Se os dois números são iguais (valor1 == valor2)
print(f"{valor1 == valor2}")

# 2. Se os dois números são diferentes (valor1 != valor2)
print(f"{valor1 != valor2}")

# 3. Se o primeiro é maior que o segundo (valor1 > valor2)
print(f"{valor1 > valor2}")

# 4. Se o primeiro é menor que o segundo (valor1 < valor2)
print(f"{valor1 < valor2}")


### 2 Completar o exercício aqui
#1
lista = ["monte", "estadio", "pista", "estrada", "estrada"]
tupla = ("monte", "estadio", "pista", "estrada", "estrada")

#2 
print("Isto é uma lista :", lista)
print("Isto é uma tupla :", tupla)

#3
print("2º elemento da lista: ", lista[1])
print("penultimo elemento da tupla: ", tupla[-2])

#4
lista[1]="novo terreno"

tupla_transformada = list(tupla)
tupla_transformada[2] = ("lago")
tupla = tuple(tupla_transformada)

#5
print("Isto é lista com 'novo terreno': ", lista)
print("Isto é tuple com 'lago': ", tupla)

#6
print("tamanho da list:", len(lista))
print("tamanho de tuple:", len(tupla))

#7
lista.append("estadio")

tupla_transformada = list(tupla)
tupla_transformada.append("pista")
tupla = tuple(tupla_transformada)

print("Lista com 'estádio' : ", lista)
print("Tuple com 'pista' :", tupla)

#8
lista.clear()

tupla_transformada = list(tupla)
tupla_transformada.clear()
tupla = tuple(tupla_transformada)

print("Isto é lista vazia: ", lista)
print("Isto é tuple vazio: ", tupla)


### 3  Completar o exercício aqui
#1
lista1 = ["monte", "estadio", "pista", "estrada", "estrada"]
um_set = set(lista1)

#2
dicionario = {"material": "pedómetro", "hidratação": "agua", "calçado":"sapatilhas","material": "pedómetro"}

#3
print("Isto é set: ", um_set)
print("Isto é dicionario: ", dicionario)

#4

dic = list(dicionario.keys())
print(dic[2])
print("Isto é dicionario sem elemento 3: ", dic)

#5 
lista1 = list(um_set)
lista1[1] = "olimpico"
print("Isto é set sem 'olimpico': ", set(lista1))

#6
print(len(um_set))
print(len(dicionario))

#7
print("estadio" in um_set)
print("material" in dicionario)


#8
um_set.add("rampa")
print("um_set: ",um_set)

dicionario["temporisador"] = "Garmin"
print("dicionario: ", dicionario)

#9
um_set.remove("pista")
dicionario.pop("material")

print("um_set: ",um_set)
print("dicionario: ", dicionario)



#### 4 Completar o exercício aqui
matriz = [ 
    [1, 1, 1],
    [2, 2, 7],
    [3, 3, 9],
    [4, 4, 13]
]

for sublista in matriz:
    elemento_somado = sum(sublista)
    sublista.append(elemento_somado)

print(matriz)

cavaleiro = { 'vida':2, 'ataque':2, 'defesa': 2, 'alcance':2 }
guerreiro = { 'vida':2, 'ataque':2, 'defesa': 2, 'alcance':2 }
arqueiro  = { 'vida':2, 'ataque':2, 'defesa': 2, 'alcance':2 }

# Completar o exercício aqui
cavaleiro = { 'vida':guerreiro['vida']*2, 'ataque':2, 'defesa': guerreiro['defesa']*2, 'alcance':2 }
guerreiro = { 'vida':2, 'ataque':cavaleiro['ataque']*2, 'defesa': 2, 'alcance':cavaleiro['alcance']*2 }
arqueiro  = { 'vida':guerreiro['vida'], 'ataque':guerreiro['ataque'], 'defesa': int(guerreiro['defesa']/2), 'alcance':guerreiro['alcance']*2 }

print(cavaleiro)
print(guerreiro)
print(arqueiro)


#### 5 Completar o exercício aqui
l=[[1,2,3],[4,5,6]]
t=([1,2,3],[4,5,6])
s={1,2,3,4,5,6}
print(type(l))
print(type(t))
print(type(s))


#### 6 Completar o exercício aqui
#1
numero_1=(input("Insira um número:"))
print("1º Numero :", numero_1)
print("Tipo de 1º Numero :", type(numero_1),"\n")

#2
numero_2=int(input("Insira outro número:"))
print("2º Numero :", numero_2)
print("Tipo de 2º Numero :", type(numero_2),"\n")

numero_3=float(input("Insira o último número:"))
print("3º Numero :", numero_3)
print("Tipo de 3º Numero :", type(numero_3),"\n")



#### 7 Completar o exercício aqui
welcome="Olá Mundo"
#1 Alinhado à direita em 20 caracteres
print(f"{'':>20},{welcome}")

#2 Trucate com indice
print(welcome[:3])

#3 Alinhamento ao centro em 20 caracteres com truncagem no segundo carácter
#print(f"{'':^20}{welcome}")
print(welcome[:1].center(20," "))

#4 format 5 int com 0 
numero_1 = 150
print('{:0>5}'.format(numero_1))

#5 format 7 int com ' '
numero_2 = 7887
print('{:' '>7}'.format(numero_2))

#6 format 3 int + 3 float
numero_3 = 20.02
numero_3 = '{:.03f}'.format(numero_3)
numero_3 = '{:0>3}'.format(numero_3)
print(numero_3)

#6 or
numero_3 = 20.02
print('{:03.3f}'.format(numero_3))


#### 8 Completar o exercício aqui
def obter_numero(mensagem):
    print(mensagem)
    numero = int(input())
    return numero

a = obter_numero("Introduz um numero:")
if a != 0:
    b = obter_numero("Introduz um numero:")
    c = obter_numero("Introduz um numero:")

    if(a<b<c):
        print("Ordem ascendente")
    elif(a>b>c):
        print("Ordem descendente")
    elif(b==0):
        if(a<c):
            print("Ordem ascendente")
        elif(a>c):
            print("Ordem descendente")
    else:
        print("Sem ordem")
        
else:
    print("Erro. O primeiro numero não pode ser 0.")



#### 9 Completar o exercício aqui
max_numero=int(input("Quantos numeros quer somar, entre 0-x? x = a valor maximo."))
iterator = 0
total= 0

for i in range(max_numero):
    total = total + i
    i=i+1
print("Total:", total)


#### 10 Completar o exercício aqui
def intro_dois_numeros_validos(mensagem):
    numero = int(input(mensagem))
    if numero > 0:
        return numero
    else:
        print("Numero não válido")

primeiro = intro_dois_numeros_validos("Introduz um numero maior que 0:")

segundo = intro_dois_numeros_validos("Introduz um numero menor que o anterior:")
if(segundo>primeiro):
    print("O segundo numero deveria ser menor que o primeiro")
else:
    print("A soma dos numeros introduzidos:", primeiro+segundo)
    print("O resto da subtração dos dois numeros introduzidos:{primeiro} - {segundo}:", primeiro-segundo)
    print("A multiplicação dos dois numeros introduzidos:", primeiro * segundo)



#### 11 Completar o exercício aqui
status = True
while status:
    numero = int(input("Introduz um numero"))
    
    if numero % 2 != 0:
        print("Numero introduzido:", numero)
    else:
        status = False


#### 12 Completar o exercício aqui
numbers = range(0,100,1)
total = sum(numbers)
print(total)



#### 13 Completar o exercício aqui
def obter_numero(mensagem):
    print(mensagem)
    numero = int(input())
    return numero

status = True
while status:
    numero_inteiro = obter_numero("Introduz um numero de 0-9")
    lista_valida = (2,4,5,7)
    if numero_inteiro in lista_valida:
        print("Numero existe na lista")
        status = False


#### 14 Completar o exercício aqui
#1
print(list(range(0,11,1)))

#2
print(list(range(-10,1,1)))

#3
print(list(range(0,22,2)))

#4
print(list(range(-19,0,1)))

#5
print(list(range(0,55,5)))



##### 15 Completar o exercício aqui
lista_1 = ["o",'l','a',' ', 'm','u','n','d','o']
lista_2 = ["o",'l','a',' ', 'l','u','a'] 
lista_3 = []

lista_3 = list(set(lista_1) - set(lista_2))
print(lista_3)


##### 16 Completar o exercício aqui
def introduzir_valor(mensagem):
    '''Para obter vários numeros'''
    numero = int(input(mensagem))
    return numero

def calcular_area(base, altura):
    '''Para calcular area'''
    resultado = base*altura
    print("A área é: ", resultado)

print("Calcular a area de um retangulo:")
base = introduzir_valor("Introduz uma base:")
altura = introduzir_valor("Introduz uma altura:")
calcular_area(base, altura)



##### 17 Completar o exercício aqui
import math
def introduzir_valor(mensagem):
    '''Para obter vários numeros'''
    numero = int(input(mensagem))
    return numero

def calcular_area(raio):
    '''Para calcular area'''
    resultado = math.pi * raio
    print("A área é: ", resultado)

raio = introduzir_valor("Introduz um raio:")
calcular_area(raio)


##### 18 Completar o exercício aqui
def obter_numero(mensagem):
    numero = int(input())
    return numero

def comparar_numeros(a,b):
    if(a>b):
        print("1")
    elif(a<b):
        print("-1")
    elif(a==b):
        print("0")

comparar_numeros(5,10)
comparar_numeros(10,5)
comparar_numeros(5,5)

primeiro = obter_numero("Introduz um numero:")
segundo = obter_numero("Introduz um numero:")
comparar_numeros(primeiro,segundo)



#### 19 Completar o exercício aqui
def ler_numero(mensagem):
    print(mensagem)
    numero = int(input())
    return numero

def maior(a,b,c):
    if(a>b and a>c):
        print("O maior numero é:", a)
    elif(b>c and b>a):
        print("O maior numero é:", b)
    elif(c>b and c>a):
        print("O maior numero é:", c)
    else:
        print("Sem ordem")

a = ler_numero("Introduz um numero:")
b = ler_numero("Introduz um numero:")
c = ler_numero("Introduz um numero:")
maior(a,b,c)

### 20 Completar o exercício aqui
def pedir_medidas(mensagem):
    print(mensagem)
    numero = float(input())
    print(numero)
    return numero

def imc(p,a):
    resultado = p/(a*a)
    if resultado < 18.50:
        return "Baixo peso"
    elif  resultado >= 18.50 and resultado < 25.00:
        return "Normal"
    elif  resultado >= 25.00 and resultado < 30.00:
        return "Sobrepeso"
    elif  resultado >= 30.00:
        return "Obesidade"
    else:
        return "Erro de calculo"


peso = pedir_medidas("Introduza o seu peso [kg]:")
altura = pedir_medidas("Introduza a sua altura [m] (max duas casas décimais): ")
print("O indice massa corporal é de", imc(peso,altura))


### 21  Fazer o exercício aqui
# Crie uma classe chamada Ponto com os atributos X e Y.
class Ponto():
    X = ''
    Y = ''

    # Adicione um construtor que inicialize X e Y com valores padrão zero.
    def __init__(self, X=0, Y=0) :
        self.X = X
        self.Y = Y

    # Implemente o método __str__ para exibir o ponto no formato (X,Y).
    def __str__(self):
        return (f"{self.X}, {self.Y}")

    def quadrante(self):
        print(self)
        if self.X == 0 and self.Y == 0 : return "Origem"
        elif self.X > 0 and self.Y > 0 : return "Primeiro Quadrante"
        elif self.X < 0 and self.Y > 0 : return "Segundo Quadrante"
        elif self.X < 0 and self.Y < 0 : return "Terceiro Quadrante"
        elif self.X > 0 and self.Y < 0 : return "Quarto Quadrante"

    # * Adicione um método vetor que recebe outro objeto ponto e calcula o vetor resultante entre os dois objetos pontos.
    def vetor(self, Ponto_B): # AB=(x2−x1,y2−y1)
        return abs(Ponto_B.X - self.X), abs(Ponto_B.Y - self.Y)

# --- Testes ---
#Criar os pontos A(2, 3), B(5,5), C(-3, -1) e D(0,0).
A = Ponto(2,3)
B = Ponto(5,5)
C = Ponto(-3,-1)
D = Ponto (0,0)

# Verifique a que quadrantes pertencem os pontos A, C e D.
print(A.quadrante())
print(B.quadrante())
print(C.quadrante())
print(D.quadrante())

# Calcule os vetores AB e BA e exiba os resultados.
print(f"Vetor AB ({B} - {A}) = {A.vetor(B)}")
print(f"Vetor BA ({A} - {B}) = {B.vetor(A)}")



### 22  Fazer o exercício aqui
# ----- IMPLEMENTAÇÃO -----
class Retangulo():
    ponto_inicial = ''
    ponto_final = ''

    def __init__(self, ponto_inicial=0, ponto_final=0):
        self.ponto_inicial = ponto_inicial
        self.ponto_final = ponto_final
    
    def base(self):
        return abs(self.ponto_inicial.X - self.ponto_final.X)

    def altura(self):
        return abs(self.ponto_inicial.X - self.ponto_final.Y)
    
    def area(self):
        return self.base() * self.altura()


# ----- EXECUÇÃO E TESTE -----
A = Ponto(2,3)
B = Ponto(5,5)
meu_retangulo = Retangulo(A,B)

print("--- Calculos de dimensões---")
base = meu_retangulo.base()
print("A base é : ", base)

altura = meu_retangulo.altura()
print("A altura é : ", altura)

area = meu_retangulo.area()
print("A área é : ", area)



### 23  Fazer o exercício aqui
# ----- IMPLEMENTAÇÃO
class Veiculo():
    def __init__(self, cor, rodas):
        self.cor = str(cor)
        self.rodas = int(rodas)
    
    def __str__(self):
        return f"\tCor: {self.cor} - \tRodas: {self.rodas}"

class Carro (Veiculo):
    def __init__(self, cor, rodas, velocidade, cilindrada):
        super().__init__(cor, rodas)
        self.velocidade = int(velocidade)
        self.cilindrada = float(cilindrada)
    
    def __str__(self):
        return f"Carro: Cor: {self.cor} \t- Rodas: {self.rodas} \t- Velocidade: {self.velocidade}km/h \t- Cilindrada: {self.cilindrada}L"

class Bicicleta (Veiculo):
    # Bicicletas não costumam ter cilindrada, mas mantive o mesmo principio, pois, as eletricas terão uma capacidade também
    def __init__(self, cor, rodas, velocidade, cilindrada=0):
        super().__init__(cor, rodas)
        self.velocidade = int(velocidade)
        self.cilindrada = float(cilindrada)
    
    def __str__(self):
        return f"Bicicleta: Cor: {self.cor} \t- Rodas: {self.rodas} \t- Velocidade: {self.velocidade}km/h \t- Cilindrada: {self.cilindrada}L"

class Moto (Veiculo):
    def __init__(self, cor, rodas, velocidade, cilindrada):
        super().__init__(cor, rodas)
        self.velocidade = int(velocidade)
        self.cilindrada = float(cilindrada)
    
    def __str__(self):
        return f"Moto: Cor: {self.cor} \t- Rodas: {self.rodas} \t- Velocidade: {self.velocidade}km/h \t- Cilindrada: {self.cilindrada}L"

class Camiao (Veiculo):
    def __init__(self, cor, rodas, velocidade, cilindrada):
        super().__init__(cor, rodas)
        self.velocidade = int(velocidade)
        self.cilindrada = float(cilindrada)
    
    def __str__(self):
        return f"Camião: Cor: {self.cor} \t- Rodas: {self.rodas} \t- Velocidade: {self.velocidade}km/h \t- Cilindrada: {self.cilindrada}L"

class Concessionario():
    def __init__(self):
        self.lista_veiculos = []

    def catalogar(self, lista_veiculos):
        self.lista_veiculos = lista_veiculos
    
    def filtrar(self, tipo_veiculo):
        """Filtra e retorna os veículos do tipo especificado da lista catalogada."""
        filtro = []
        for veiculo in self.lista_veiculos:
            if type(veiculo).__name__ == tipo_veiculo:
                filtro.append(veiculo) 
        return filtro


# ----- EXECUÇÃO -----
lista_veiculos=[]

Carro1 = Carro("vermelho", 4, 180, 2.4)
Carro2 = Carro("azul", 4, 150, 1.4)
print(Carro1)
lista_veiculos.append(Carro1)
lista_veiculos.append(Carro2)

Bicicleta1 = Bicicleta("preto", 2, 20) # Não precisa de cilindrada
Bicicleta2 = Bicicleta("preto", 2, 40)
print(Bicicleta1)
lista_veiculos.append(Bicicleta1)
lista_veiculos.append(Bicicleta2)

Moto1 = Moto("branco", 2, 180, 0.8)
Moto2 = Moto("preto", 2, 240, 1.2)
print(Moto1)
lista_veiculos.append(Moto1)
lista_veiculos.append(Moto2)

Camiao1 = Camiao("vermelho", 8, 90, 15)
Camiao2 = Camiao("amarelo", 8, 110, 20)
print(Camiao1)
lista_veiculos.append(Camiao1)
lista_veiculos.append(Camiao2)


# ----- TESTE -----
print("\n--- Catálogo e Filtro ---")
Concessionario_CarrosUsados = Concessionario()
Concessionario_CarrosUsados.catalogar(lista_veiculos)
filtro = Concessionario_CarrosUsados.filtrar("Carro") 

print(f"Veículos encontrados ({len(filtro)}):")
for veiculo in filtro:
    print(f"- {veiculo}")


### 24  Fazer o exercício aqui
class GerenciadorLista:
    def __init__(self, lista_inicial=None):
        if lista_inicial is None:
            self.lista = []
        else:
            self.lista = list(lista_inicial)
        
    def adicionar_elemento(self, elemento):
        try: 
            print(f"\nA tentar adicionar: {elemento}")
            if elemento not in self.lista:
                self.lista.append(elemento)
                print(f"Elemento '{elemento}' adicionado com sucesso.")
                print(f"Conteúdo atual da lista: {self.lista}")
            else:
                raise ValueError(f"Erro: Impossível adicionar elementos duplicados => [{elemento}].")
        except:
                print(f"Item duplicated : {elemento}. Não adicionado.")

# --- Implementação ---
gerenciador = GerenciadorLista()
print(f"Lista Inicial: {gerenciador.lista}")
gerenciador.adicionar_elemento(10)
gerenciador.adicionar_elemento(-2)
gerenciador.adicionar_elemento("Olá")

# --- Teste de erro ---
gerenciador.adicionar_elemento("Olá")



#### 25 Fazer o exercício aqui
from datetime import datetime, timedelta
#import pytz

# 1 Apresentar datas  atuais
today = datetime.now()
print("datetimwe now:", today)

# lisboa_now = pytz.timezone('Europe/Lisbon')
# print(lisboa_now.tz.zone, end=" ")

# 2 Formatos de datas
print("\ndata e hora em formatos diferentes:")
print("{}/{}/{}".format(today.day, today.month, today.year))
print("{}-{}-{} {}:{}".format(today.year, today.month, today.day, today.hour, today.minute))

# 3 Duração Projetos
print("\nCalcular duração projeto:")
try:
    start_date = input("Data de inicio do projeto?") # Exemplo: 2025-11-09 23:31
    end_date = input("Date de termino do projeto?") # Exemplo: 2025-12-19 23:31
except ValueError:
    print("\nERRO: Formato de data inválido!")
    print(f"Certifica-te que usas o formato: 2025-12-19 23:31")
formato = "%Y-%m-%d %H:%M" # => CORRECÇÃO DE: delta = end_date + start_date # reage como con
start = datetime.strptime(end_date, formato) # => É preciso separar, pois neste formato não 
end = datetime.strptime(start_date, formato)
delta = start - end
print(delta)

# 4 Calcular DeadLines
print("\nCalcular a data final do vencimento do projeto:")
tempo_tarefa = int(input("Dias de tempo de tarefa?")) # Exemplo: 5 (dias)
t = timedelta(days = int(tempo_tarefa), hours=0, seconds = 0000)
vencimento_date = today + t
print(f"{vencimento_date.year}-{vencimento_date.month}-{vencimento_date.day}") # => Igual re


#### 26 Fazer o exercício aqui
from datetime import datetime
class Evento():
    def __init__(self, titulo, data_hora_inicio, data_hora_fim):
        formato = "%Y-%m-%d %H:%M:%S"
        self.titulo = titulo
        self.data_hora_inicio = datetime.strptime(data_hora_inicio, formato) # Importante que esta conversão seja feita logo com o construtor para que possa ser usa
        self.data_hora_fim = datetime.strptime(data_hora_fim, formato)
    
    def __str__(self):
        return f"Evento: {self.titulo} \t| {self.data_hora_inicio} \t| {self.data_hora_fim}" 


class Agenda():
    def __init__(self):
        self.lista_eventos = []
    
    def adicionar_evento(self, evento):
        self.lista_eventos.append(evento)


    def listar_eventos(self):
        '''Lista todos os eventos na agenda, mostrando o título, a data e hora de início, e a data e hora de término.'''
        for evento in self.lista_eventos:
            print(f"{evento.titulo} - {evento.data_hora_inicio} - {evento.data_hora_fim}")


    def proximo_evento(self):
        '''Mostra o próximo evento na agenda com base na data atual "M4_01_Outras ferramentas_de_POO.ipynb"'''
        data_atual = datetime.now()
        for evento in self.lista_eventos:
            if  evento.data_hora_inicio > data_atual:
                print(f"{evento.titulo} - {evento.data_hora_inicio} - {evento.data_hora_fim}")


    def eventos_do_dia(self, data):
        '''Eventos_do_dia. Detalhe:
        - datetime: É uma data e hora específica (Ex: 25 de Dezembro às 10:00). Não tem .days.
        - timedelta: É um intervalo de tempo (Ex: 5 dias de diferença). Este é que tem .days.'''
        data_pesquisa = datetime.strptime(data, "%Y-%m-%d").date()
        encontrou = False
        if len(self.lista_eventos)> 0:
            for evento in self.lista_eventos:
                if evento.data_hora_inicio.date() == data_pesquisa:
                    print(f"{evento.titulo} - {evento.data_hora_inicio} - {evento.data_hora_fim}")
                    encontrou = True
        if not encontrou:
            print(f"Não existem eventos nesta data")

# 1 Crie um objeto da classe Agenda.
agenda = Agenda()

# 2 Crie pelo menos 5 objetos da classe Evento e adicione à agenda.
agenda.adicionar_evento(Evento("Ler", "2025-02-09 00:22:01", "2025-12-20 23:59:01"))
agenda.adicionar_evento(Evento("Trabalhar", "2025-12-11 00:22:58", "2025-12-27 23:59:02"))
agenda.adicionar_evento(Evento("Almoçar", "2026-12-14 00:22:57", "2025-12-18 23:59:03"))
agenda.adicionar_evento(Evento("Jantar", "2026-12-14 00:22:57", "2025-12-18 23:59:03"))
agenda.adicionar_evento(Evento("Estudar", "2026-12-23 00:22:56", "2025-12-19 23:59:04"))
agenda.adicionar_evento(Evento("Conferencia","2025-12-30 00:22:55", "2025-12-22 23:59:05"))

# 3 Liste todos os eventos.
print("\n3\) Listar eventos:")
agenda.listar_eventos()

# 4 Mostre o próximo evento.
print("\n4\) Mostrar proximo evento:")
agenda.proximo_evento()

# 5 Liste todos os eventos para uma data específica
print("\n5\) Listar eventos de uma data especifica:")
agenda.eventos_do_dia("2025-02-08")
agenda.eventos_do_dia("2026-12-12")
agenda.eventos_do_dia("2026-12-14")



#### 27 Fazer o exercício aqui
class DivisaoPorZeroHandler():
    def __init__(self, denominador, numerador):
        self.denominador = denominador
        self.numerador = numerador
    
    def calcular(self):
        try:
            resultado = self.numerador / self.denominador
            print(resultado)
        
        except ZeroDivisionError:
            print("Não se pode dividir um valor por zero '0'. Escolhe um numero maior")

calculoOk = DivisaoPorZeroHandler(1,5)
calculoOk.calcular()

calculoNok = DivisaoPorZeroHandler(0,3)
calculoNok.calcular()



#### 28 Fazer o exercício aqui
class IndiceForaDosLimitesHandler():
    def __init__(self,lista):
        self.lista = lista
    
    def aceder_elemento(self, indice):
        try:
            elemento = self.lista[indice]
            print(elemento)
        except IndexError:
            print(f"O índice fornecido ({indice}) está fora do intervalo válido. O comprimento da lista é {len(self.lista)}.")
            print("Essa posição não existe. Tera de escolher um valor diferente")

lista = [0,3,6,9,12]
valor_lista = IndiceForaDosLimitesHandler(lista)
valor_lista.aceder_elemento(2)
valor_lista.aceder_elemento(6)


class GerenciadorLista:
    def __init__(self, lista_inicial=None):
        if lista_inicial is None:
            self.lista = []
        else:
            self.lista = list(lista_inicial)
        
    def adicionar_elemento(self, elemento):
        #print(f"\nA tentar adicionar: {elemento}")
        try:
            if elemento in self.lista:
                # Lançamos o erro para que quem chama o método decida o que fazer
                #print(f"Item duplicated : {elemento}. Não adicionado.")
                raise ValueError(f"Impossível adicionar elementos duplicados => [{elemento}].")
            else:
                self.lista.append(elemento)
            #print(f"Elemento '{elemento}' adicionado com sucesso.")
            print(f"Conteúdo atual da lista: {self.lista}")
        except ValueError as e:
            print(f"Erro: {e}")

# --- Implementação ---
gerenciador = GerenciadorLista()
print(f"Lista Inicial: {gerenciador.lista}")
gerenciador.adicionar_elemento(10)
gerenciador.adicionar_elemento(-2)
gerenciador.adicionar_elemento("Olá")

# --- Teste do erro com lista vazia ---
print(" ----- ")
gerenciador.adicionar_elemento("Olá")
gerenciador.adicionar_elemento(-2)

# --- Teste do erro com lista já preenchida ---
print(" ----- ")
gerenciador = GerenciadorLista([10])
print(f"Lista Inicial: {gerenciador.lista}")
gerenciador.adicionar_elemento(10)



### 29 Fazer o exercício aqui
# ---- Funções ---
def criar_ficheiro(nome_ficheiro, conteudo):
    print("criar ficheiro de raiz, para o exercicio: ", nome_ficheiro)
    file = open(nome_ficheiro, "w", encoding="UTF8") 
    file.write(conteudo)
    file.close()

def ler_ficheiro(nome_do_ficheiro):
    lista_final = []
    with open(nome_do_ficheiro, "r", encoding="UTF8") as ficheiro:
        for linha in ficheiro:
            dados_pessoa = linha.strip().split(",") # .strip() remove o \n invisível no fim da linha
            #linha = (f"nome:{dados_pessoa[0]}, idade:{dados_pessoa[1]}, email:{dados_pessoa[2]}")
            lista_final.append(dados_pessoa)
        return lista_final

def calcular_media_idades(conteudo):
    total_pessoas = 0
    idades = 0
    for pessoa in conteudo:
        if int(pessoa[1]) > 0:
            idades += int(pessoa[1])
            total_pessoas += 1
    
    media = idades / total_pessoas
    print(media)

def adicionar_pessoa(nome_do_ficheiro, nome, idade, mail):
    if (len(str(nome)) >= 5) and (int(idade) > 0):
        linha = f"'],[{nome},{idade},{mail}\n"
        # print(linha)

        try:
            with open(nome_do_ficheiro, "a", encoding="UTF8") as file:
                file.write(linha)
                print(f"Dados para pessoa {nome}, foram adicionados ao ficheiro.")
            
        except Exception as e:
            print(f"Erro ao escrever no ficheiro: {e}")
    else:
        print(f"Dados inválidos. É necessáro verificar o nome '{str(nome)}' e / ou a idade'{idade}'.")

# --- Implementação --
# - Criar ficheiro
nome_do_ficheiro = "dados.txt"
texto_ficheiro_novo = "João,25,joao@example.com\nMaria,30,maria@example.com\nCarlos,22,carlos@example.com\nAna,27,ana@example.com"
criar_ficheiro(nome_do_ficheiro, texto_ficheiro_novo)

# - Ler e Mostrar resultado
conteudo = ler_ficheiro(nome_do_ficheiro)
print(conteudo)

# - Calcular media de idades
calcular_media_idades(conteudo)

# - Adicionar nova pessoa
adicionar_pessoa(nome_do_ficheiro, "MarioCapela", "37", "mario@gmail.com")
adicionar_pessoa(nome_do_ficheiro, "test", "-1", "teste@gmail.com")

# - Confirmar controlo de nome e de idade adicionado
conteudo = ler_ficheiro(nome_do_ficheiro)
print(conteudo)


#### 30

# Fazer o exercício aqui
# ---- Classes
class Produto():
    """"Representa as propriedades de um produto"""
    def __init__(self, nome, preco, quantidade):
        self.nome = str(nome)
        self.preco = int(preco)
        self.quantidade = int(quantidade)
    
class GerenciadorDeProdutos():
    """"Representa as propriedades de um fiheiro com produtos"""
    def __init__(self, nome_ficheiro):
        self.nome_ficheiro = nome_ficheiro
        file = open(nome_ficheiro, "a", encoding="UTF8") 
        file.write("")
        file.close()

    def adicionar_produto(self, produto):
        detalhes = (f"{produto.nome},{produto.quantidade},{produto.preco}\n")
        #with open(self.nome_ficheiro, "a", encoding="UTF8") as file:
        #    file.write(detalhes)

        file = open(nome_ficheiro, "a", encoding="UTF8") 
        file.write(detalhes)
        file.close()

    def retornar_todos_os_dados(self): # print=> (nome=Arroz) - [QUT=15] | {preço=34}
        with open(self.nome_ficheiro, "r", encoding="UTF8") as file:
            for conteudo in file:
                detalhes = conteudo.replace('\n', '').split(",")
                # o método replace corrige o espaço final \n que existe, caso contrario, o output seria:
                # (nome=Arroz) - [QUT=34] | {preço=15
                # }"
                print(f"(nome={detalhes[0]}) - [QUT={detalhes[1]}] | {{preço={detalhes[2]}}}")

        """A abordagem seguinte não funciona. além disso, como não lê linha à linha, impossibilita o mesmo tratamento de dados.
        Output: 
            (nome=Arroz) - [QUT=34] | {preço=15
            Carne} """
        #file = open(self.nome_ficheiro, "r", encoding="UTF8")
        #produtos = file.read()
        #file.close()
        #detalhes = produtos.strip().split(",")
        #print(f"(nome={detalhes[0]}) - [QUT={detalhes[1]}] | {{preço={detalhes[2]}}}")
    
    def apagar_conteudo(self):
        """Criação extra enunciado, para que o ficheiro possa ser reiniciado em testes"""
        file = open(self.nome_ficheiro, "w", encoding="UTF8") 
        file.write("")
        file.close()

# ---- Implementação
# - Criar 5 objectos
produto1 = Produto("Arroz", 15,34) 
produto2 = Produto("Carne", 2,20)
produto3 = Produto("Tomate", 4,4)
produto4 = Produto("Laranja", 8,3)
produto5 = Produto("Peixe", 6,30)

# - Teste
# print(produto1.nome,produto2.nome, produto3.nome, produto4.nome, produto5.nome )



#### 31 - Criar Ficheiro
nome_ficheiro = "dados_produto.txt"
gerenciador = GerenciadorDeProdutos(nome_ficheiro)

# - Escrever objetos no ficheiro
gerenciador.adicionar_produto(produto1)
gerenciador.adicionar_produto(produto2)
gerenciador.adicionar_produto(produto3)
gerenciador.adicionar_produto(produto4)
gerenciador.adicionar_produto(produto5)

# - Retornar dados (a interpretação feita a partir do enunciado,
# foi que a leitura do ficheiro e a sua impressão poderiam ser realizadas na mesma função)
gerenciador.retornar_todos_os_dados()

# - Para reiniciar o ficheiro
#gerenciador.apagar_conteudo()


#### 32  EXERCICIO EXTRAS
'''
### EXERCÍCIO 1
Neste exercício,devemoscriarumprogramaqueabra e leia um ficheiro de texto e nos proporcione a seguinte
informação:
• Nome doficheiro
• Se o ficheiro está fechado ou não (True ouFalse)
• O modo de abertura doficheiro
O ficheirodetextovai denominar-seprovas.txt e teráoseguinte conteúdoemtextoplano(criarpreviamente):

Olá, que tal está
Isto é uma prova
Hojenãochove
Aproxima-seo natal
Até logoe muitoboas

Pesquisar nadocumentação quemétodosproporcionama informaçãonecessária.Por último,listaroconteúdo doficheiro.
Aviso importante: se quiser ler um ficheiro escrito diretamente com Python, então é possível que encontre erros de codificação ao mostrar
alguns caracteres. Assegurar-se que indica a codificação do ficheiro manualmente durante a abertura como argumento no open, por
exemplo com UTF-8: open(..., encoding="utf8")
'''
print("\n ------- \nExercicio 1")

def criar_ficheiro(nome_ficheiro, conteudo):
    print("criar ficheiro de raiz, para o exercicio")
    file = open(nome_ficheiro, "w", encoding="UTF8") 
    file.write(conteudo)
    file.close()

text="Olá, que tal está\nIsto é uma prova\nHojenãochove\nAproxima-se o natal\nAté logo e muitoboas"
criar_ficheiro("provas.txt", text)



'''
EXERCÍCIO 2
Sobreoficheiroanterior, vamos obter a seguinte informação:
• Leitura de uma linhado ficheiro
• Leitura do ficheiro linha a linha
'''
print("\n ------- \nExercicio 2")

# Leitura do ficheiro
print("\nLeitura do ficheiro")
file = open("provas.txt", "r", encoding="UTF8")
texto = file.read()
file.close()
print(texto)

# Leitura de todo o ficheiro mas numa linha apenas
print("\nLeitura de todo o ficheiro mas numa linha apenas")
file = open("provas.txt", "r", encoding="UTF8")
texto = file.readlines()
file.close()
print(texto)

# Leitura de uma linha do ficheiro
print("\nLeitura de uma linha do ficheiro")
file = open("provas.txt", "r", encoding="UTF8")
texto = file.readline()
file.close()
print(texto)

# Leitura do ficheiro linha a linha
print("\nLeitura do ficheiro linha a linha")
with open("provas.txt", "r", encoding="UTF8") as ficheiro:
    for linha in ficheiro:
        print(linha)



'''
EXERCÍCIO 3
Sobre o ficheiro anterior,fazer a seguinte operação:
• Abrir o ficheiro em modo escrita (ligar o conteúdo, sem o substituir) e escrever uma nova linha de texto.
Rever se está a funcionar corretamente comprovandoo ficheiro (fechar e abrir para poder visualizar as alterações).
'''
print("\n ------- \nExercicio 3")
def confirmar_escrita_ficheiro(nome_ficheiro):
    print("\nNome do ficheiro: ",nome_ficheiro)
    file = open(nome_ficheiro, "r", encoding="UTF8")
    texto = file.read()
    file.close()
    print(texto)

def adicionar_novo_texto(nome_ficheiro, nova_linha):
    file = open(nome_ficheiro, "a", encoding="UTF8") # a = append
    file.write(nova_linha)
    file.close()

adicionar_novo_texto("provas.txt", ", esta é uma nova linha")
confirmar_escrita_ficheiro("provas.txt")


'''
EXERCÍCIO 4
Sobre o ficheiro anterior,fazer a seguinte operação:
• Abrir o ficheiro em modo escrita (substituindo o conteúdo) e escrever uma nova linha de texto.
Rever se está a funcionar corretamente comprovando o ficheiro (fechar e abrir para poder visualizar as alterações).
'''

print("\n ------- \nExercicio 4")
def substituir_texto_no_ficheiro(nome_ficheiro, novo_texto):
    file = open(nome_ficheiro, "w", encoding="UTF8") # w = write
    file.write(novo_texto)
    file.close()

substituir_texto_no_ficheiro("provas.txt", "substituir texto escrito no ficheiro por algo novo")
confirmar_escrita_ficheiro("provas.txt")



'''
EXERCÍCIO 5
Neste exercício, devemos criar um programa para ler os dado de um ficheiro de texto, que transforme cada fila num
dicionário e o adiciona a uma lista chamada pessoas. Logo, recorre-se às pessoas da lista e para cada uma mostrar
deforma amigável todos os seus campos.Por exemplo,assim:
(id=1) Carlos Pérez =>05/01/1989
O ficheiro de texto denominar pessoas.txt e terá o seguinte conteúdo no texto plano(criá-lo previamente):
1;Carlos;Pérez;05/01/1989
2;Manuel;Heredia;26/12/1973
3;Rosa;Campos;12/06/1961
4;David;García;25/07/2006

Os campos do dicionário serão por ordem:id, nome, apelido e nascimento.
Aviso importante: se quiser ler um ficheiro que não se escreveu diretamente com o Python, então é possível que se encontrem erros de
codificação ao mostrar alguns caracteres.
Assegurar que indica a codificação do ficheiro manualmente durante a abertura como argumento no open, por exemplo com UTF-8:
open(..., encoding="utf8")
Pista: é possível que se tenha que recorrer a funções como replace ou split para poder modificar o texto de uma linha de texto. Pesquisar
documentação sobre isso.
'''
print("\n ------- \nExercicio 5")
'''Usando as funções criadas anteriormente, este exercicio fica da seguinte forma'''

nome_do_ficheiro = "pessoas.txt"
texto_ficheiro_novo = "1;Carlos;Pérez;05/01/1989\n2;Manuel;Heredia;26/12/1973\n3;Rosa;Campos;12/06/1961\n4;David;García;25/07/2006"
criar_ficheiro(nome_do_ficheiro, texto_ficheiro_novo)
confirmar_escrita_ficheiro(nome_do_ficheiro)

pessoa = []
with open(nome_do_ficheiro, "r", encoding="UTF8") as ficheiro:
    for linha in ficheiro:
        pessoa = linha.split(";")
        print(f"(id={pessoa[0]}) {pessoa[1]} {pessoa[2]} =>{pessoa[3]}") # (id=1) Carlos Pérez =>05/01/1989





#### 33
import os, random
print("\nClassificação:")

class Corredor:
    def __init__(self, posicao, nome, nacionalidade, equipa, pontos):
        self.posicao = posicao
        self.nome = nome 
        self.nacionalidade = nacionalidade 
        self.equipa = equipa 
        self.pontos = int(pontos)

class Classificacao:
    def __init__(self, nome_ficheiro):
        self.nome_ficheiro = nome_ficheiro
        # Cria o ficheiro com o cabeçalho apenas se ele não existir
        if not os.path.exists(self.nome_ficheiro):
            with open(self.nome_ficheiro, "w", encoding="UTF8") as file:
                file.write("Posicao,Nome,Nacionalidade,Equipa,Points\n")
    
    def reiniciar_ficheiro(self):
        """Apaga tudo e escreve apenas o cabeçalho para evitar erros de leitura"""
        with open(self.nome_ficheiro, "w", encoding="UTF8") as file:
            file.write("Posicao,Nome,Nacionalidade,Equipa,Pontos\n")
        
    def adicionar_jogador(self, jogador):
        try:
            linha = f"{jogador.posicao},{jogador.nome},{jogador.nacionalidade},{jogador.equipa},{jogador.pontos}\n"
            with open(self.nome_ficheiro, "a", encoding="UTF8") as file:
                file.write(linha)
        except Exception as e:
            print(f"Erro ao gravar: {e}")

    def mostrar_classificacao(self):
        print(f"\n{'POS':<4} | {'PILOTO':<20} | {'NAC':<5} | {'EQUIPA':<20} | {'PTS'}")
        print("-" * 65)
        try:
            with open(self.nome_ficheiro, "r", encoding="UTF8") as file:
                next(file) # Pula a primeira linha (cabeçalho)
                for conteudo in file:
                    detalhes = conteudo.strip().split(",")
                    if len(detalhes) == 5:
                        print(f"{detalhes[0]:<4} | {detalhes[1]:<20} | {detalhes[2]:<5} | {detalhes[3]:<20} | {detalhes[4]}")
        except FileNotFoundError:
            print("Ficheiro não encontrado.")

    @staticmethod
    def gerar_vencedores_aleatorios():
        pilotos = [
            "Lando Norris", "Max Verstappen", "Oscar Piastri", "George Russell", 
            "Charles Leclerc", "Lewis Hamilton", "Kimi Antonelli", "Alexander Albon", 
            "Carlos Sainz", "Fernando Alonso", "Nico Hulkenberg", "Isack Hadjar", 
            "Oliver Bearman", "Liam Lawson", "Esteban Ocon", "Lance Stroll", 
            "Yuki Tsunoda", "Pierre Gasly", "Gabriel Bortoleto", "Franco Colapinto",
            "Jack Doohan"
        ]
        return random.sample(pilotos, 10)

    def atribuir_pontos_corrida(self, vencedores):
        pontuacao = [25, 18, 15, 12, 10, 8, 6, 4, 2, 1]
        dados_pilotos = []

        with open(self.nome_ficheiro, "r", encoding="UTF8") as file:
            linhas = file.readlines()
            header = linhas[0]
            
            for linha in linhas[1:]:
                dados = linha.strip().split(",")
                nome_piloto = dados[1]
                pontos_atuais = int(dados[4])

                if nome_piloto in vencedores:
                    indice_vitoria = vencedores.index(nome_piloto)
                    pontos_atuais += pontuacao[indice_vitoria]

                piloto = f"{dados[0]},{dados[1]},{dados[2]},{dados[3]},{pontos_atuais}\n"
                dados_pilotos.append(piloto)
                dados_pilotos.sort(key=lambda x: x[4], reverse=True)

        with open(self.nome_ficheiro, "w", encoding="UTF8") as file:
            file.write(header)
            file.writelines(dados_pilotos)
        
    def ordenar_por_pontos(self):
        try:
            dados_pilotos = []
            with open(self.nome_ficheiro, "r", encoding="UTF8") as file:
                linhas = file.readlines()
                header = linhas[0]
                for linha in linhas[1:]:
                    dados = linha.strip().split(",")
                    if len(dados) == 5:
                        dados[4] = int(dados[4]) 
                        dados_pilotos.append(dados)
            dados_pilotos.sort(key=lambda x: x[4], reverse=True)
            linhas_ordenadas = [header]
            for i, piloto in enumerate(dados_pilotos, 1):
                nova_linha = f"{i},{piloto[1]},{piloto[2]},{piloto[3]},{piloto[4]}\n"
                linhas_ordenadas.append(nova_linha)

            with open(self.nome_ficheiro, "w", encoding="UTF8") as file:
                file.writelines(linhas_ordenadas)
            print("✨ Classificação e posições atualizadas:")
        except Exception as e:
            print(f"Erro ao ordenar: {e}")

# --- INSTANCIAÇÃO ---
documento = "M5_01_ExercicioExtra_Formula1.txt"
grand_premio = Classificacao(documento)

# Limpar o ficheiro antes de começar a adicionar os jogadores
grand_premio.reiniciar_ficheiro()

jogadores = [
    Corredor("1", "Lando Norris", "GBR", "McLaren", "0"),
    Corredor("2", "Max Verstappen", "NED", "Red Bull Racing", "0"),
    Corredor("3", "Oscar Piastri", "AUS", "McLaren", "0"),
    Corredor("4", "George Russell", "GBR", "Mercedes", "0"),
    Corredor("5", "Charles Leclerc", "MON", "Ferrari", "0"),
    Corredor("6", "Lewis Hamilton", "GBR", "Ferrari", "0"),
    Corredor("7", "Kimi Antonelli", "ITA", "Mercedes", "0"),
    Corredor("8", "Alexander Albon", "THA", "Williams", "0"),
    Corredor("9", "Carlos Sainz", "ESP", "Williams", "0"),
    Corredor("10", "Fernando Alonso", "ESP", "Aston Martin", "0"),
    Corredor("11", "Nico Hulkenberg", "GER", "Kick Sauber", "0"),
    Corredor("12", "Isack Hadjar", "FRA", "Racing Bulls", "0"),
    Corredor("13", "Oliver Bearman", "GBR", "Haas F1 Team", "0"),
    Corredor("14", "Liam Lawson", "NZL", "Racing Bulls", "0"),
    Corredor("15", "Esteban Ocon", "FRA", "Haas F1 Team", "0"),
    Corredor("16", "Lance Stroll", "CAN", "Aston Martin", "0"),
    Corredor("17", "Yuki Tsunoda", "JPN", "Red Bull Racing", "0"),
    Corredor("18", "Pierre Gasly", "FRA", "Alpine", "0"),
    Corredor("19", "Gabriel Bortoleto", "BRA", "Kick Sauber", "0"),
    Corredor("20", "Franco Colapinto", "ARG", "Alpine", "0"),
    Corredor("21", "Jack Doohan", "AUS", "Alpine", "0")
]

# Adicionar todos os jogadores ao ficheiro
for jogador in jogadores:
    grand_premio.adicionar_jogador(jogador)

# Mostrar resultados
grand_premio.mostrar_classificacao()

# Simular 4 Provas => atribuir classificações
for i in range(0, 4):
    vencedores_da_prova = Classificacao.gerar_vencedores_aleatorios()
    print(f"\n\n🏆 RESULTADO PROVA {i+1}: {vencedores_da_prova[0]} venceu!")
    grand_premio.atribuir_pontos_corrida(vencedores_da_prova)
    grand_premio.ordenar_por_pontos()
    grand_premio.mostrar_classificacao()


#####
'''
#### Criação
* Crie um ficheiro Excel chamado __"dados.xlsx"__ com os dados fornecidos a seguir, utilizando a biblioteca __openpyxl__:

| Nome      | Idade | Nota Matemática | Nota Português | Nota Ciências |
|-----------|-------|-----------------|----------------|---------------|
| João      | 16    | 8               | 7              | 6             |
| Maria     | 17    | 19              | 12             | 14            |
| Pedro     | 16    | 6               | 9              | 10            |
| Ana       | 18    | 15              | 20             | 17          
* Adicione uma nova coluna chamada "Avaliação" em cada disciplina (Matemática, Português e Ciências).
* A coluna Avaliação deve conter o valor "Suficiente" se a nota for igual ou superior a 10, ou "Insuficiente" se for inferior a 10.
* Crie uma nova folha chamada "Resumo", que deve incluir:
    * O número total de alunos.
    * A quantidade de notas suficientes e insuficientes por disciplina (Matemática, Português, Ciências).
* Salve o ficheiro modificado como __"dados_analise.xlsx"__.caracteres).
'''


# 34 Fazer o exercício aqui
import openpyxl

# A. Criar um ficheiro
def criar_configurar_ficheiro(nome):
    tabela = openpyxl.Workbook()
    
    # Selecionar a folha ativa e mudar o título
    planilha = tabela.active 
    planilha.title = "planilha"
    
    # 1. Definir o cabeçalho usando uma lista simples
    cabecalho = ["Nome", "Idade", "Nota Matemática", "Nota Português", "Nota Ciências"]
    planilha.append(cabecalho) # O append adiciona a lista inteira na primeira linha disponível

    # 2. Dados em formato de lista de strings (como no seu exemplo)
    conteudo = [
        "João, 16, 8, 7, 6",
        "Maria, 17, 19, 12, 14",
        "Pedro, 16, 6, 9, 10",
        "Ana, 18, 15, 20, 17"
    ]

    # 3. Adicionar conteúdo percorrendo a lista
    for linha_texto in conteudo:
        dados_separados = linha_texto.split(", ")   # Transformar "João, 16, ..." em ["João", "16", ...]
        planilha.append(dados_separados)             # Adicionar a linha tratada à planilha
    
    # 4. Guardar o ficheiro
    tabela.save(nome)
    print(f"Ficheiro {nome} criado com sucesso!")
    tabela.close()

def mostrar_tabela(nome_ficheiro):
    wb = openpyxl.load_workbook(nome_ficheiro)
    folha_um = wb['planilha']

    # Imprime por colunas
    for linhas in folha_um.columns:
        for columna in linhas:
            print(columna.coordinate, columna.value)
        print("------------")
    wb.close()


# B. Adicione uma nova coluna chamada "Avaliação" em cada disciplina (Matemática, Português e Ciências).
def rescrever_tabela_avaliacao(nome):
    tabela = openpyxl.Workbook()
    planilha = tabela.active 
    planilha.title = "planilha"
    
    cabecalho = ["Nome", "Idade", "Nota Matemática","Avaliação Matemática", "Nota Português", "Avaliação Português", "Nota Ciências", "Avaliação Ciências"]
    planilha.append(cabecalho) # O append adiciona a lista inteira na primeira linha disponível

    conteudo = [
        "João, 16, 8, 7, 6",
        "Maria, 17, 19, 12, 14",
        "Pedro, 16, 6, 9, 10",
        "Ana, 18, 15, 20, 17"
    ]

    for linha_texto in conteudo:
        dados_separados = linha_texto.split(", ")
        # print(dados_separados) # Para verificar o que está dentro antes de entrar
        linha_completa =[
            dados_separados[0],                        # primeira coluna => Nome
            dados_separados[1],                        # segundo coluna => Idade
            dados_separados[2],                        # terceira coluna => Matemática
            avaliacao_da_nota(dados_separados[2], 10), # quarta coluna => Validado
            dados_separados[3],                        # quinta coluna => Português
            avaliacao_da_nota(dados_separados[3], 10), # sexta coluna => Validado
            dados_separados[4],                        # séptima coluna => Ciências
            avaliacao_da_nota(dados_separados[4], 10), # oitava coluna => Validado
        ]
        # print(linha_completa) # Para verificar o que está dentro quando sai
        planilha.append(linha_completa)    
    
    tabela.save(nome)
    print(f"\nFicheiro {nome} construido com sucesso!")
    tabela.close()

def avaliacao_da_nota(dados, nota_suficiente):
    ''' Função que auxilia a transformação de dados de forma simplificada '''
    return "Suficiente" if int(dados) >= nota_suficiente else "Insuficiente"



# C. Crie uma nova folha chamada "Resumo" :* O número total de alunos.* A quantidade de notas suficientes e insuficientes por disciplina (Matemática, Português, Ciências).
def criar_nova_folha(nome_ficheiro):
    wb = openpyxl.load_workbook(nome_ficheiro)
    folha_um = wb[wb.sheetnames[0]]
    folha_dois = wb.create_sheet("Resumo")
    # print(wb.sheetnames) # Para verificar
    
    folha_dois["A1"] = "Numero total de alunos"
    folha_dois.append(["Disciplina","Suficiente","Insuficiente"])

    total_alunos = 0
    disciplina = {} # dicionáirio

    for coluna in folha_um.columns:
        titulo_coluna = coluna[0].value # cabeçalho da coluna
        if titulo_coluna == "Nome":
            for i in range(1, len(coluna)):
                total_alunos += 1
                # print("alunos", total_alunos)

        elif "Avaliação" in titulo_coluna:
            suficiente = 0
            insuficiente = 0
            for i in range(1, len(coluna)):
                # print("coluna valor", coluna[i].value)
                if str(coluna[i].value) == "Suficiente":
                    suficiente += 1
                elif str(coluna[i].value) == "Insuficiente":
                    insuficiente += 1
            disciplina[titulo_coluna] = [suficiente, insuficiente]
    # print(disciplina) # Para verificar

    folha_dois["B1"] = total_alunos
    for disciplina, classificacao in disciplina.items():
        folha_dois.append([disciplina, classificacao[0],classificacao[1]])

    # D. Salve o ficheiro modificado como "dados_analise.xlsx".caracteres).
    wb.save(nome_ficheiro)
    nome_ficheiro = (nome_ficheiro+".caracteres")
    wb.save(nome_ficheiro)
    wb.close()

def total_avaliacao_disciplina(disciplina, coluna):
    ''' Serve para contar os Suficientes e Insuficientes por cada disciplina, devolvendo uma lista [suficiente][insuficiente]/disciplina'''
    suficiente = 0
    insuficiente = 0
    if str(coluna.value) == "Suficiente":
        suficiente += 1
    if str(coluna.value) == "Insuficiente":
        insuficiente += 1
    print("suficiente", suficiente)
    print("insuficiente", insuficiente)
    return [[suficiente], [insuficiente]]

nome_ficheiro = "dados.xlsx"
criar_configurar_ficheiro(nome_ficheiro)
# mostrar_tabela(nome_ficheiro) # Para verificar
rescrever_tabela_avaliacao(nome_ficheiro)
# mostrar_tabela(nome_ficheiro) # Para verificar
criar_nova_folha(nome_ficheiro)
# mostrar_tabela(nome_ficheiro) # Para verificar


############# 
## 35
'''
#### Criação

* Crie o ficheiro __"dados.csv"__ com os dados fornecidos, utilizando a biblioteca csv:

| Produto   | Categoria   | Preço  | Quantidade |
|-----------|-------------|--------|------------|
| Teclado   | Informática | 20.00  | 100        |
| Mouse     | Informática | 15.00  | 200        |
| Cadeira   | Móveis      | 100.00 | 50         |
| Caneca    | Utensílios  | 5.00   | 300        |
| Mesa      | Móveis      | 150.00 | 20         |
| Cadeira   | Móveis      | 250.00 | 20         |

* Ordene os dados pela coluna "Preço", do maior para o menor.
* Filtrar os dados por produtos com preço acima de 25.
* Após ordenar e filtrar, salve os resultados em um novo ficheiro CSV chamado __"dados_filtrados.csv"__.
'''

import csv

# A. Criar um ficheiro
def criar_configurar_ficheiro(nome):
    # Criar documento
    with open(nome, "w", encoding="utf-8", newline='') as tabela:
        tabela = open(nome,"w")
        
        # 2. Dados em formato de lista de strings
        conteudo = [
            "Teclado, Informatica, 20.00, 100",
            "Mouse, Informatica, 15.00, 200",
            "Cadeira, Móveis, 100.00, 50",
            "Caneca, Utensílios, 5.00, 300",
            "Mesa, Móveis, 150.00, 20",
            "Cadeira, Móveis, 250.00, 20"
        ]

    # 1. Adicionar conteúdo percorrendo a lista
    writer = csv.writer(tabela)
    writer.writerow(['Produto', 'Categoria','Preço','Quantidade'])
    for linha in conteudo:
        limpos = linha.split(", ") # .split(", ") divide a string em colunas reais
        writer.writerow(limpos)

    # 2. Guardar o ficheiro
    #del writer
    tabela.close()
    print(f"Ficheiro {nome} criado com sucesso!")

def mostrar_tabela(nome_ficheiro):
    with open(nome_ficheiro, "r") as ficheiro:
        reader = csv.reader(ficheiro, delimiter=',')
        
        for line, row in enumerate(reader):
            next(reader)
            print('Linha: ' + str({line}))
            print(f"Produto: {row[0]} | Categoria: {row[1]} | Preço: {row[2]} | Quantidade': {row[3]}")
        print("------------")
    ficheiro.close()

def ordenar_tabela(nome_ficheiro):
    '''Para Ordernar a tabela'''
    print("\n --- Tabela ordenada por preço ---")
    with open(nome_ficheiro, "r") as ficheiro:
        # DictReader para podermos usar o nome da coluna 'Preço'
        reader = csv.DictReader(ficheiro)
        lista_dados = list(reader)
        
    # Ordenar convertendo a string para float para ordenação numérica
    lista_final = sorted(lista_dados, key=lambda x: float(x['Preço']), reverse=False)

    for row in enumerate(lista_final):
        print(f"Produto: {row}")
    print("------------")
    ficheiro.close()

def filtrar_tabela_guardar(nome_ficheiro, nome_ficheiro_final, preco): 
    '''Para filtrar os dados sobre base de um valor'''
    count_files=0

    # Para ler, guardar e fechar automaticamente
    with open(nome_ficheiro,"r") as ficheiro:
        reader = csv.DictReader(ficheiro)

        # Continua a tratar de projetos de forma progressivas
        with open(nome_ficheiro_final, mode='w', encoding='utf-8', newline='') as outfile:
            writer = csv.DictWriter(outfile, fieldnames=reader.fieldnames)
            writer.writeheader()
            
            for row in reader:
                # print("row:", row)# to teste
                if float(row['Preço']) >= preco: # Filter 
                    count_files += 1
                    writer.writerow(row)

    print(f"Tabela filtrada com :", count_files, "entradas")

nome_ficheiro = "dados.csv"
nome_ficheiro_final = "dados_filtrados.csv"

criar_configurar_ficheiro(nome_ficheiro)
mostrar_tabela(nome_ficheiro) # Para verificar
ordenar_tabela(nome_ficheiro)
filtrar_tabela_guardar(nome_ficheiro, nome_ficheiro_final, 25.00)
#mostrar_tabela(nome_ficheiro_final) # Para verificar


######
## 36
'''
#### Criação

* Crie o ficheiro Excel "vendas.xlsx" com os dados fornecidos a seguir:

| Mês        | Vendas  |
|------------|---------|
| Janeiro    | 2000    |
| Fevereiro  | 3000    |
| Março      | 2500    |
| Abril      | 4000    |
| Maio       | 3500    |
| Junho      | 4500    |

* Utilize a biblioteca __openpyxl__ para gerar um gráfico de linha que mostre a evolução das vendas ao longo dos meses.
* Adicione o gráfico à folha, posicionando-o abaixo da tabela de vendas.
* Salve o ficheiro como __"vendas_com_grafico.xlsx"__.
'''

# Fazer o exercício aqui
import openpyxl
from openpyxl.chart import AreaChart, LineChart, Reference
#from openpyxl.chart.axis import DateAxis

# A. Criar um ficheiro
def criar_configurar_ficheiro(nome):
    tabela = openpyxl.Workbook()
    
    # Selecionar a folha ativa e mudar o título
    planilha = tabela.active 
    planilha.title = "Relatório de Vendas"
    
    # 1. Definir o cabeçalho usando uma lista simples
    cabecalho = ["Mês", "Vendas"]
    planilha.append(cabecalho) # O append adiciona a lista inteira na primeira linha disponível

    # 2. Dados em formato de lista de strings (como no seu exemplo)
    conteudo = [
        "Janeiro, 2000",
        "Fevereiro, 3000",
        "Março, 2500",
        "Abril, 4000",
        "Maio, 3500",
        "Junho, 4500"
    ]

    # 3. Adicionar conteúdo percorrendo a lista
    for linha_texto in conteudo:
        dados_separados = linha_texto.split(", ")
        planilha.append(dados_separados) 
    
    # 4. Guardar o ficheiro
    tabela.save(nome)
    print(f"Ficheiro {nome} criado com sucesso!")
    tabela.close()

def mostrar_tabela(nome_ficheiro):
    wb = openpyxl.load_workbook(nome_ficheiro)
    folha_um = wb['Relatório de Vendas']

    # Imprime por colunas
    for linhas in folha_um.columns:
        for columna in linhas:
            print(columna.coordinate, columna.value)
        print("------------")
    wb.close()

def gerar_grafico_linhas(nome_ficheiro, novo_ficheiro_final):
    wb = openpyxl.load_workbook(nome_ficheiro)
    folha = wb["Relatório de Vendas"]

    # Criar o objeto de gráfico de linha
    grafico = LineChart() # AreaChart() 
    grafico.title = "Evolução de Vendas"
    grafico.style = 13
    
    # Definir os dados (vendas) e as categorias (meses)
    grafico.x_axis.title = 'Meses'
    grafico.y_axis.title = 'Vendas'

    # Os dados estão na coluna A1:B7
    meses = Reference(folha, min_col=1, min_row=2, max_row=7) # Categorias => A1: A7 (meses)
    grafico.set_categories(meses)
    
    data = Reference(folha, min_col=2, min_row=1, max_col=2, max_row=7) # Informação => B1: B7 (vendas)
    grafico.add_data(data, titles_from_data=True, from_rows=False)
    
    # Estilo do gráfico
    grafico.y_axis.scaling.min = 0
    grafico.y_axis.scaling.max = 6000
    grafico.y_axis.majorUnit = 500
    grafico.series[0].smooth = True
    
    # 5. Adicionar o gráfico à planilha abaixo da tabela (com 7 linhas)
    folha.add_chart(grafico, "A9")
    
    # 6. Salvar
    wb.save(novo_ficheiro_final)
    print("Gráfico criado com sucesso no arquivo:" + novo_ficheiro_final)

nome_ficheiro = "vendas.xlsx"
novo_ficheiro_final = "vendas_com_grafico.xlsx"

criar_configurar_ficheiro(nome_ficheiro)
mostrar_tabela(nome_ficheiro) # Para verificar
gerar_grafico_linhas(nome_ficheiro, novo_ficheiro_final)


# 37 Fazer o exercício aqui

'''
#### Criação

* Crie o JSON a seguir como uma string em Python:

{
'marcadores': [{'latitude': 40.416875,
   'longitude': -3.703308,
   'city': 'Madrid',
   'description': 'Puerta del Sol'},
  {'latitude': 40.417438,
   'longitude': -3.693363,
   'city': 'Madrid',
   'description': 'Paseo del Prado'},
  {'latitude': 40.407015,
   'longitude': -3.691163,
   'city': 'Madrid',
   'description': 'Estación de Atocha'}]
}

* Utilize  o módulo __json__ para carregar o conteúdo dessa string e transformá-lo em um dicionário Python.
* Exiba os dados de maneira legível somente das descrições de cada ponto turístico.
* Salve o JSON as informações no ficheiro "__marcadores.json__", como: do nome da cidade, latitude e longitude.
'''
import json

def transformar_em_dicionario(lista):
    print(lista)
    lista_dictionary = json.loads(lista)
    # print(lista_dictionary) # Resultado para testes
    return lista_dictionary

def guardar_ficheiro(dicionario, ficheiro_nome):
    with open(ficheiro_nome, "w", encoding="utf-8") as ficheiro:
        json.dump(dicionario, ficheiro, indent=4, ensure_ascii=False) # Transformar de python para JSON, Gravando o objeto diretamente no ficheiro

def ler_ficheiro(ficheiro_nome):
    with open(ficheiro_nome, "r", encoding="utf-8") as ficheiro: # Lê o conteúdo do ficheiro e Transformar em dicionário
        dados = json.load(ficheiro) 
        
        frutas = dados["Frutaria"]["Frutas"]
        print("\nQuantidades de frutas:")
        for fruta in frutas:
            print(f"{fruta['Nome']}: {fruta['Quantidade']}")

        legumes = dados["Frutaria"]["Legumes"]
        print("\nQuantidades de legumes:")
        for legume in legumes:
            print(f"{legume['Nome']}: {legume['Quantidade']}")

mercearia = '{"Frutaria": { "Frutas": [{"Nome":"Bananas", "Quantidade":5}, {"Nome":"Mangas", "Quantidade":10}, {"Nome":"Maracuja", "Quantidade":6}], "Legumes": [{"Nome":"Batata Doce", "Quantidade":10}, {"Nome":"Corgete", "Quantidade":2}, {"Nome":"Beterraba", "Quantidade":4}]}}'
ficheiro_nome = "frutaria.json"
# print("Mercearia",type(mercearia)) #class 'str' => OK

dicionario = transformar_em_dicionario(mercearia)
# print(type(dicionario))  #class 'dict' => OK

guardar_ficheiro(dicionario, ficheiro_nome)
ler_ficheiro(ficheiro_nome)


# 38
'''
[
    {"nome": "João", "idade": 28, "interesses": ["música", "futebol", "cinema"]},
    {"nome": "Maria", "idade": 34, "interesses": ["literatura", "viajar", "pintura"]},
    {"nome": "Ana", "idade": 22, "interesses": ["fotografia", "moda", "tecnologia"]}
]

* Converta esse conteúdo em um ficheiro CSV com a seguinte estrutura:

| Nome  | Idade | Interesse 1   | Interesse 2   | Interesse 3    |
|-------|-------|---------------|---------------|----------------|
| João  | 28    | música        | futebol       | cinema         |
| Maria | 34    | literatura    | viajar        | pintura        |
| Ana   | 22    | fotografia    | moda          | tecnologia      |

* Salve o ficheiro CSV com o nome "__usuarios.csv__".
* Depois faça a leitura do conteúdo do ficheiro "__usuarios.csv__" para certificar que foi salvo corretamente.x"__.
'''
import json
# {string} em dicionario = {"":""} + loads
# dicionario em json => dump
# json en dicionario => load

def criar_ficheiro_json(nome_ficheiro_final, dados_para_transformar):
    # Para que os dados possam ter estrutura, tem de se substituir pelas aspas corretas
    dados_corrigidos = dados_para_transformar.replace("'", '"')
    # Transformar string em dicionario
    dicionario = json.loads(dados_corrigidos) # load a "S" => loads

    with open(nome_ficheiro_final, 'w', encoding='utf-8') as f:
    # Transformar dicionario em JSON
        json.dump(dicionario, f, indent=4)

# Exiba os dados de maneira legível somente das descrições de cada ponto turístico.
def mostrar_dados(nome_ficheiro_final, valor_a_imprimir):
    with open(nome_ficheiro_final, 'r', encoding='utf-8') as ficheiro:
        cidades = json.load(ficheiro)
    
    marcadores_ficheiro = []
    dados = ()
    # Aproveitar para que ao ler/ imprimir, se faça uma triagem imediadta. Assim, não se tem de percorrer novamente a lista
    for cidade in cidades['marcadores']:
        dados = {'city': cidade['city'],'latitude': cidade['latitude'],'longitude': cidade['longitude']}
        marcadores_ficheiro.append(dados)
        print(cidade[valor_a_imprimir]) # Imprime as descrições
    
    # print("dados_ficheiro", dados_ficheiro)
    guardar_json(nome_ficheiro_final, marcadores_ficheiro)

# Salve o JSON as informações no ficheiro "marcadores.json", como: do nome da cidade, latitude e longitude.
def guardar_json(nome_ficheiro_final, dados):
    json_marcadores = {'marcadores': dados}
    # print(json_marcadores) # Para testar
    with open(nome_ficheiro_final, 'w', encoding='utf-8') as ficheiro:
        json.dump(json_marcadores, ficheiro, indent=4,) 

# Crie o JSON a seguir como uma string em Python:
dados_para_transfomar = "{'marcadores': [{'latitude': 40.416875,'longitude': -3.703308,'city': 'Madrid','description': 'Puerta del Sol'},{'latitude': 40.417438,'longitude': -3.693363,'city': 'Madrid','description': 'Paseo del Prado'},{'latitude': 40.407015,'longitude': -3.691163,'city': 'Madrid','description': 'Estación de Atocha'}]}"
nome_ficheiro_final = "marcadores.json"

# Utilize  o módulo json para carregar o conteúdo dessa string e transformá-lo em um dicionário Python.
criar_ficheiro_json(nome_ficheiro_final,dados_para_transfomar)
mostrar_dados(nome_ficheiro_final, "description")



# 39
import json,csv

# Carregue o JSON a seguir em uma lista de dicionários. 
def carregar_dicionario(json_pessoas):
    return json.loads(json_pessoas)

# Converta esse conteúdo em um ficheiro CSV na estrutura
def converter_para_csv(dicionario, nome_ficheiro):
    cabecalho = ["Nome", "Idade", "Interesse 1", "Interesse 2", "Interesse 3"]

    with open(nome_ficheiro, "w", encoding="utf-8", newline='') as tabela:
        tabela = open(nome_ficheiro,"w")

    # Adicionar conteúdo percorrendo a lista
    writer = csv.writer(tabela)
    writer.writerow(cabecalho)

    for pessoa in dicionario:
        dados = [pessoa['nome'], pessoa['idade']]
        dados.extend(pessoa['interesses']) # para ler cada item, sem ter de percorrer com um novo 'for'
        writer.writerow(dados)

    #tabela.save(nome_ficheiro)
    print(f"Ficheiro {nome_ficheiro} criado com sucesso!")
    tabela.close()


# Depois faça a leitura do conteúdo do ficheiro "usuarios.csv"
def imprimir(ficheiro):
    with open(nome_ficheiro, "r") as ficheiro:
        reader = csv.reader(ficheiro)
        # Imprime por colunas
        for linha in reader:
            print(" | ".join(linha))

json_pessoas='[{"nome": "João", "idade": 28, "interesses": ["música", "futebol", "cinema"]},{"nome": "Maria", "idade": 34, "interesses": ["literatura", "viajar", "pintura"]},{"nome": "Ana", "idade": 22, "interesses": ["fotografia", "moda", "tecnologia"]}]'
nome_ficheiro ="usuarios.csv"

dicionario = carregar_dicionario(json_pessoas)
converter_para_csv(dicionario, nome_ficheiro)
imprimir(nome_ficheiro)



########### 41
import json

def transformar_em_dicionario(lista):
    print(lista)
    lista_dictionary = json.loads(lista)
    # print(lista_dictionary) # Resultado para testes
    return lista_dictionary

def guardar_ficheiro(dicionario, ficheiro_nome):
    with open(ficheiro_nome, "w", encoding="utf-8") as ficheiro:
        json.dump(dicionario, ficheiro, indent=4, ensure_ascii=False) # Transformar de python para JSON, Gravando o objeto diretamente no ficheiro

def ler_ficheiro(ficheiro_nome):
    with open(ficheiro_nome, "r", encoding="utf-8") as ficheiro: # Lê o conteúdo do ficheiro e Transformar em dicionário
        dados = json.load(ficheiro) 
        
        frutas = dados["Frutaria"]["Frutas"]
        print("\nQuantidades de frutas:")
        for fruta in frutas:
            print(f"{fruta['Nome']}: {fruta['Quantidade']}")

        legumes = dados["Frutaria"]["Legumes"]
        print("\nQuantidades de legumes:")
        for legume in legumes:
            print(f"{legume['Nome']}: {legume['Quantidade']}")

mercearia = '{"Frutaria": { "Frutas": [{"Nome":"Bananas", "Quantidade":5}, {"Nome":"Mangas", "Quantidade":10}, {"Nome":"Maracuja", "Quantidade":6}], "Legumes": [{"Nome":"Batata Doce", "Quantidade":10}, {"Nome":"Corgete", "Quantidade":2}, {"Nome":"Beterraba", "Quantidade":4}]}}'
ficheiro_nome = "frutaria.json"
# print("Mercearia",type(mercearia)) #class 'str' => OK

dicionario = transformar_em_dicionario(mercearia)
# print(type(dicionario))  #class 'dict' => OK

guardar_ficheiro(dicionario, ficheiro_nome)
ler_ficheiro(ficheiro_nome)



############ 42
import json
# {string} em dicionario = {"":""} + loads
# dicionario em json => dump
# json en dicionario => load

def criar_ficheiro_json(nome_ficheiro_final, dados_para_transformar):
    # Para que os dados possam ter estrutura, tem de se substituir pelas aspas corretas
    dados_corrigidos = dados_para_transformar.replace("'", '"')
    # Transformar string em dicionario
    dicionario = json.loads(dados_corrigidos) # load a "S" => loads

    with open(nome_ficheiro_final, 'w', encoding='utf-8') as f:
    # Transformar dicionario em JSON
        json.dump(dicionario, f, indent=4)

# Exiba os dados de maneira legível somente das descrições de cada ponto turístico.
def mostrar_dados(nome_ficheiro_final, valor_a_imprimir):
    with open(nome_ficheiro_final, 'r', encoding='utf-8') as ficheiro:
        cidades = json.load(ficheiro)
    
    marcadores_ficheiro = []
    dados = ()
    # Aproveitar para que ao ler/ imprimir, se faça uma triagem imediadta. Assim, não se tem de percorrer novamente a lista
    for cidade in cidades['marcadores']:
        dados = {'city': cidade['city'],'latitude': cidade['latitude'],'longitude': cidade['longitude']}
        marcadores_ficheiro.append(dados)
        print(cidade[valor_a_imprimir]) # Imprime as descrições
    
    # print("dados_ficheiro", dados_ficheiro)
    guardar_json(nome_ficheiro_final, marcadores_ficheiro)

# Salve o JSON as informações no ficheiro "marcadores.json", como: do nome da cidade, latitude e longitude.
def guardar_json(nome_ficheiro_final, dados):
    json_marcadores = {'marcadores': dados}
    # print(json_marcadores) # Para testar
    with open(nome_ficheiro_final, 'w', encoding='utf-8') as ficheiro:
        json.dump(json_marcadores, ficheiro, indent=4,) 

# Crie o JSON a seguir como uma string em Python:
dados_para_transfomar = "{'marcadores': [{'latitude': 40.416875,'longitude': -3.703308,'city': 'Madrid','description': 'Puerta del Sol'},{'latitude': 40.417438,'longitude': -3.693363,'city': 'Madrid','description': 'Paseo del Prado'},{'latitude': 40.407015,'longitude': -3.691163,'city': 'Madrid','description': 'Estación de Atocha'}]}"
nome_ficheiro_final = "marcadores.json"

# Utilize  o módulo json para carregar o conteúdo dessa string e transformá-lo em um dicionário Python.
criar_ficheiro_json(nome_ficheiro_final,dados_para_transfomar)
mostrar_dados(nome_ficheiro_final, "description")



################ 43 

import json
import csv

# Carregue o JSON a seguir em uma lista de dicionários. 
def carregar_dicionario(json_pessoas):
    return json.loads(json_pessoas)

# Converter o conteúdo em um ficheiro CSV na estrutura
def converter_para_csv(dicionario, nome_ficheiro):
    cabecalho = ["Nome", "Idade", "Interesse 1", "Interesse 2", "Interesse 3"]

    with open(nome_ficheiro, "w", newline='') as tabela: # '' => é para que ao abrir o arquivo, sejam evitadas linhas em branco
        writer = csv.writer(tabela)
        writer.writerow(cabecalho)

        for pessoa in dicionario:
            dados = [pessoa['nome'], pessoa['idade']]
            dados.extend(pessoa['interesses']) # para ler cada item, sem ter de percorrer com um novo 'for'
            writer.writerow(dados)

    print(f"Ficheiro {nome_ficheiro} criado com sucesso!")
    tabela.close()


# Fazer a leitura do conteúdo do ficheiro "usuarios.csv"
def imprimir(nome_ficheiro):
    """Lê o CSV e imprime-o formatado como uma tabela alinhada no terminal."""
    try:
        with open(nome_ficheiro, "r") as ficheiro:
            reader = csv.reader(ficheiro)
            
            # Formatação tabelas e colunas
            print(f"{'TABELA DE USUÁRIOS':^75}")
            print("-" * 75)
            
            for i, linha in enumerate(reader):
                print(f"{linha[0]:<12} | {linha[1]:<6} | {linha[2]:<15} | {linha[3]:<15} | {linha[4]:<15}")

                if i == 0: # Linha decorativa após o cabeçalho
                    print("-" * 75)

    except FileNotFoundError:
        print("Erro: O ficheiro não existe.")

json_pessoas='[{"nome": "João", "idade": 28, "interesses": ["música", "futebol", "cinema"]},{"nome": "Maria", "idade": 34, "interesses": ["literatura", "viajar", "pintura"]},{"nome": "Ana", "idade": 22, "interesses": ["fotografia", "moda", "tecnologia"]}]'
nome_ficheiro ="usuarios.csv"

dicionario = carregar_dicionario(json_pessoas)
converter_para_csv(dicionario, nome_ficheiro)
imprimir(nome_ficheiro)